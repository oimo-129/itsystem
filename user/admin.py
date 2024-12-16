from django.contrib import admin  
from django.contrib.auth.admin import UserAdmin  
from django.contrib.auth.forms import UserChangeForm, UserCreationForm  
from .models import UserProfile, Department,NeedForm

#添加用户请求管理模块
from openpyxl import Workbook
from django.http import HttpResponse
#处理url编码
from urllib.parse import quote
#加上时间
from datetime import datetime
#处理excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 自定义用户修改表单  
class CustomUserChangeForm(UserChangeForm):  
    class Meta(UserChangeForm.Meta):  
        model = UserProfile  

# 自定义用户创建表单  
class CustomUserCreationForm(UserCreationForm):  
    class Meta(UserCreationForm.Meta):  
        model = UserProfile  

# 自定义用户管理类  
class CustomUserAdmin(UserAdmin):  
    form = CustomUserChangeForm  
    add_form = CustomUserCreationForm  
    
    # 列表显示字段  
    list_display = ('username', 'email', 'desp_name', 'is_active', 'is_staff', 'date_joined')  
    
    # 列表筛选字段  
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'desp_name')  
    
    # 搜索字段  
    search_fields = ('username', 'email')  
    
    # 排序字段  
    ordering = ('-date_joined',)  
    
    # 编辑页面字段分组  
    fieldsets = (  
        (None, {'fields': ('username', 'password')}),  
        ('个人信息', {'fields': ('email', 'department', 'avatar')}),  
        ('权限', {  
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),  
        }),  
        ('重要日期', {'fields': ('last_login', 'date_joined')}),  
    )  
    
    # 添加用户时的字段分组  
    add_fieldsets = (  
        (None, {  
            'classes': ('wide',),  
            'fields': ('username', 'email', 'desp_name', 'password1', 'password2'),  
        }),  
    )  
    
    # 添加批量操作  
    actions = ['reset_password']  
    
    def reset_password(self, request, queryset):  
        """批量重置密码为 '123456' """  
        from django.contrib.auth.hashers import make_password  
        default_password = '123456'  
        
        for user in queryset:  
            user.password = make_password(default_password)  
            user.save()  
            
        self.message_user(request, f'已成功重置 {queryset.count()} 个用户的密码为: {default_password}')  
    reset_password.short_description = '重置所选用户密码为123456'  

# 注册部门模型  
class DepartmentAdmin(admin.ModelAdmin):  
    list_display = ['name']  
    search_fields = ['name']  
    
    
    
    
    
    
    
    
    
    
    
    
    
#解释器语法     
  
@admin.register(NeedForm)
class NeedFormAdmin(admin.ModelAdmin):
    list_display = ['user', 'submit_time', 'section', 'analysis_type', 'urgency', 'department', 'contact_person']
    list_filter = ['urgency', 'submit_time', 'department']
    search_fields = ['user__username', 'content', 'section']
    
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        
        # 设置文件名
        current_time = datetime.now().strftime('%Y年%m月%d日')
        filename = f'家电一院情报需求_{current_time}.xlsx'
        encoded_filename = quote(filename)
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "需求表单"
        
        # 定义表头
        columns = [
            '用户名', '提交时间', '涉及板块', '分析大类', '分析细类',
            '紧急程度', '具体内容', '需求科室', '对接人员', '联系方式'
        ]
        
        # 设置列宽（根据需要调整数值）
        column_widths = {
            '用户名': 15,
            '提交时间': 20,
            '涉及板块': 15,
            '分析大类': 15,
            '分析细类': 15,
            '紧急程度': 25,
            '具体内容': 40,
            '需求科室': 15,
            '对接人员': 15,
            '联系方式': 20,
        }
        
        # 写入表头并设置样式
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            
            # 设置表头样式
            cell.font = Font(
                name='微软雅黑',
                bold=True,
                size=11,
                color='000000'  # 黑色
            )
            cell.fill = PatternFill(
                start_color='CCE5FF',  # 浅蓝色背景
                end_color='CCE5FF',
                fill_type='solid'
            )
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列宽
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = column_widths[column_title]
        
        # 写入数据并设置样式
        for row_num, obj in enumerate(queryset, 2):
            row = [
                obj.user.username,
                obj.submit_time.strftime('%Y-%m-%d %H:%M'),
                obj.section,
                obj.analysis_type,
                obj.analysis_detail,
                obj.get_urgency_display(),
                obj.content,
                obj.department,
                obj.contact_person,
                obj.contact_info
            ]
            
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                
                # 设置数据单元格样式
                cell.font = Font(name='微软雅黑', size=10)
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='center',
                    wrap_text=True
                )
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 冻结表头
        ws.freeze_panes = 'A2'
        
        wb.save(response)
        return response
    
    export_as_excel.short_description = '导出选中项到Excel'

# 注册模型到 admin  
admin.site.register(UserProfile, CustomUserAdmin)  
admin.site.register(Department, DepartmentAdmin)
# 修改管理后台的标题和头部  
admin.site.site_header = '家电一院情报系统后台'  # 设置网站页头  
admin.site.site_title = '家电一院情报系统'   # 设置页面标题  
admin.site.index_title = '管理'  # 设置首页标题